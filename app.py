import streamlit as st
import pandas as pd
import openpyxl
import io
import numpy as np
from datetime import datetime
from zoneinfo import ZoneInfo

# 1. CONFIGURATION DE LA PAGE WEB
st.set_page_config(page_title="Générateur de Rapport", layout="wide")

if 'version_compteur' not in st.session_state:
    st.session_state.version_compteur = 1

st.title(f"Générateur de Rapport : Rabais Fournisseurs (v{st.session_state.version_compteur})")

montreal_time = datetime.now(ZoneInfo("America/Montreal")).strftime("%Y-%m-%d à %H:%M:%S")
st.caption(f"📅 Génération du rapport — Heure de Montréal : {montreal_time}")

st.write("Veuillez téléverser votre fichier d'inventaire brut ci-dessous.")

# 2. BOUTON D'IMPORTATION
fichier_upload = st.file_uploader("Choisissez le fichier de commandes (.xlsx)", type=["xlsx"])

if fichier_upload is not None:
    st.info(f"Traitement instantané et rigoureux (Version {st.session_state.version_compteur}) en cours...")
    
    try:
        wb = openpyxl.load_workbook(fichier_upload, data_only=False)
        
        if 'Rabais fournisseurs' in wb.sheetnames and 'Rabais entre 2 dates' in wb.sheetnames:
            ws_cmd = wb['Rabais fournisseurs']
            
            df_cmd = pd.read_excel(fichier_upload, sheet_name='Rabais fournisseurs')
            df_rabais = pd.read_excel(fichier_upload, sheet_name='Rabais entre 2 dates')
            
            df_cmd.columns = df_cmd.columns.str.strip()
            df_rabais.columns = df_rabais.columns.str.strip()
            
            max_row = ws_cmd.max_row
            tolerance = 10
            
            # Nettoyage et types
            df_cmd['Qté_commandée'] = pd.to_numeric(df_cmd['Qté_commandée'], errors='coerce').fillna(0)
            df_cmd['Montant_ST'] = pd.to_numeric(df_cmd['Montant_ST'], errors='coerce').fillna(0)
            df_cmd['Date_Facture'] = pd.to_datetime(df_cmd['Date_Facture'], errors='coerce')
            
            # --- IDENTIFICATION DES COLONNES PAR LEUR NOM ---
            col_cred = 'Clé_unique_détail_credité' if 'Clé_unique_détail_credité' in df_cmd.columns else ([c for c in df_cmd.columns if 'crédit' in str(c).lower() and 'clé' in str(c).lower()][0] if any('crédit' in str(c).lower() and 'clé' in str(c).lower() for c in df_cmd.columns) else None)
            col_date_recl_cred = 'Date_réclamé_détail_credité' if 'Date_réclamé_détail_credité' in df_cmd.columns else ([c for c in df_cmd.columns if 'crédit' in str(c).lower() and 'date' in str(c).lower()][0] if any('crédit' in str(c).lower() and 'date' in str(c).lower() for c in df_cmd.columns) else None)
            
            col_promo = None
            for candidate in ['Code_de_Promotion', 'Code_promotion', 'Code promotion', 'Code de Promotion']:
                if candidate in df_cmd.columns:
                    col_promo = candidate
                    break
            if not col_promo:
                col_promo = next((c for c in df_cmd.columns if 'promo' in str(c).lower()), None)
            
            # --- 1. PRÉ-CALCUL RAPIDE DES RABAIS ENTRE 2 DATES ---
            prod_col = '# Produit' if '# Produit' in df_rabais.columns else df_rabais.columns[1]
            debut_col = 'Date début' if 'Date début' in df_rabais.columns else [c for c in df_rabais.columns if 'début' in str(c).lower()][0]
            fin_col = 'Date échéance' if 'Date échéance' in df_rabais.columns else [c for c in df_rabais.columns if 'échéance' in str(c).lower() or 'fin' in str(c).lower()][0]
            rabais_col = 'Rabais' if 'Rabais' in df_rabais.columns else [c for c in df_rabais.columns if 'rabais' in str(c).lower()][0]
            
            df_rabais[debut_col] = pd.to_datetime(df_rabais[debut_col], errors='coerce')
            df_rabais[fin_col] = pd.to_datetime(df_rabais[fin_col], errors='coerce')
            rabais_dict = {prod: group for prod, group in df_rabais.groupby(prod_col)}
            
            rabais_calc_arr = np.zeros(len(df_cmd))
            date_deb_arr = [''] * len(df_cmd)
            date_fin_arr = [''] * len(df_cmd)
            
            for idx, row in df_cmd.iterrows():
                prod = row.get('No_Produit', None)
                df_date = row.get('Date_Facture', pd.NaT)
                qte = row.get('Qté_commandée', 0)
                
                if prod in rabais_dict and pd.notnull(df_date):
                    sub = rabais_dict[prod]
                    valid = sub[(sub[debut_col] <= df_date + pd.Timedelta(days=tolerance)) & (sub[fin_col] >= df_date - pd.Timedelta(days=tolerance))].copy()
                    if not valid.empty:
                        valid['dH'] = (valid[debut_col] - df_date).abs()
                        valid['dI'] = (valid[fin_col] - df_date).abs()
                        valid['dist'] = valid[['dH', 'dI']].min(axis=1)
                        valid = valid.sort_values(by=['dist', debut_col], ascending=[True, False])
                        best = valid.iloc[0]
                        taux = float(best[rabais_col]) if pd.notnull(best[rabais_col]) else 0
                        rabais_calc_arr[idx] = taux * qte
                        date_deb_arr[idx] = best[debut_col].strftime('%Y-%m-%d') if pd.notnull(best[debut_col]) else ''
                        date_fin_arr[idx] = best[fin_col].strftime('%Y-%m-%d') if pd.notnull(best[fin_col]) else ''
            
            df_cmd['_rabais_calc'] = rabais_calc_arr

            # --- 2. PRÉ-CALCUL COLONNE N ---
            col_n_arr = np.zeros(len(df_cmd))
            for _, group in df_cmd.groupby('Clé_unique_détail_commande'):
                max_r = group['_rabais_calc'].max()
                col_n_arr[group[group['_rabais_calc'] == max_r].index] = 1
            df_cmd['_col_N'] = col_n_arr

            # --- 3. PRÉ-CALCUL GLOBAL DES ENSEMBLES ET CONDITIONS ---
            cles_reclamees = set(df_cmd[df_cmd['Date_Réclamée'].notnull() & (df_cmd['Date_Réclamée'].astype(str).str.strip() != '') & (df_cmd['Date_Réclamée'].astype(str) != 'NaT')]['Clé_unique_détail_commande'].dropna().astype(str))
            cles_facture = set(df_cmd['Clé_unique_détail_facture'].dropna().astype(str)) if 'Clé_unique_détail_facture' in df_cmd.columns else set()
            cles_credite = set(df_cmd[col_cred].dropna().astype(str)) if col_cred else set()

            group_keys = ['Clé_unique_détail_commande', 'No_Produit']
            sum_qte_group = df_cmd.groupby(group_keys)['Qté_commandée'].transform('sum')
            
            has_date_rc = pd.Series(False, index=df_cmd.index)
            if col_date_recl_cred and col_date_recl_cred in df_cmd.columns:
                d_col = df_cmd[col_date_recl_cred]
                has_date_rc = d_col.notnull() & (d_col.astype(str).str.strip() != "") & (d_col.astype(str) != "NaT") & (d_col.astype(str) != "nan")
            group_has_date_rc = has_date_rc.groupby(list(df_cmd[group_keys].columns)).transform('any')

            has_valid_cred = pd.Series(False, index=df_cmd.index)
            if col_cred and col_cred in df_cmd.columns:
                c_col = df_cmd[col_cred]
                has_valid_cred = c_col.notnull() & (c_col.astype(str).str.strip() != "") & (c_col.astype(str) != "0") & (c_col.astype(str) != "nan")
            group_cred_count = has_valid_cred.groupby(list(df_cmd[group_keys].columns)).transform('sum')

            mask_no_recl_series = df_cmd['Date_Réclamée'].isnull() | (df_cmd['Date_Réclamée'].astype(str).str.strip() == "") | (df_cmd['Date_Réclamée'].astype(str) == "NaT")
            df_cmd['_temp_i_rab'] = np.where(mask_no_recl_series, df_cmd['_rabais_calc'], -np.inf)
            max_rab_i_group = df_cmd.groupby(group_keys)['_temp_i_rab'].transform('max')

            max_fact_n1_filled = pd.Series(np.nan, index=df_cmd.index)
            if 'Clé_unique_détail_facture' in df_cmd.columns:
                n1_mask = df_cmd['_col_N'] == 1
                df_cmd['_max_fact_n1'] = np.where(n1_mask, df_cmd['Clé_unique_détail_facture'], np.nan)
                max_fact_n1_filled = df_cmd.groupby('Clé_unique_détail_commande')['_max_fact_n1'].transform('max')

            # --- 4. INDEXATION POUR RECHERCHEX DE L'ONGLET 'Rabais entre 2 dates' ---
            # Colonnes H (début), I (fin), B (produit), L (code promotion) dans df_rabais
            h_col = df_rabais.columns[7] if len(df_rabais.columns) > 7 else [c for c in df_rabais.columns if 'début' in str(c).lower()][0]
            i_col = df_rabais.columns[8] if len(df_rabais.columns) > 8 else [c for c in df_rabais.columns if 'échéance' in str(c).lower() or 'fin' in str(c).lower()][0]
            b_col = df_rabais.columns[1] if len(df_rabais.columns) > 1 else prod_col
            l_col = df_rabais.columns[11] if len(df_rabais.columns) > 11 else [c for c in df_rabais.columns if 'promo' in str(c).lower()][0]

            rabais_lookup = {}
            for _, r_row in df_rabais.iterrows():
                val_h = pd.to_datetime(r_row.get(h_col), errors='coerce')
                val_i = pd.to_datetime(r_row.get(i_col), errors='coerce')
                val_b = str(r_row.get(b_col, '')).strip()
                val_l = r_row.get(l_col, '')
                
                if pd.notnull(val_h) and pd.notnull(val_i):
                    str_h = val_h.strftime('%Y-%m-%d')
                    str_i = val_i.strftime('%Y-%m-%d')
                    key_x = f"{str_h}{str_i}{val_b}"
                    rabais_lookup[key_x] = val_l

            # --- 5. ÉCRITURE INSTANTANÉE DANS EXCEL ---
            for r in range(2, max_row + 1):
                idx = r - 2
                if idx < len(df_cmd):
                    row = df_cmd.loc[idx]
                    
                    qte = row['Qté_commandée']
                    montant_st = row['Montant_ST']
                    
                    # Valeurs de date début (Col S / 19), date fin (Col T / 20) et No_Produit ou equivalent (Col AB / 28)
                    # Dans notre script, date_deb_arr[idx] et date_fin_arr[idx] contiennent les dates S et T formatées, et row.get('No_Produit') correspond à AB
                    d_s = date_deb_arr[idx]
                    d_t = date_fin_arr[idx]
                    val_ab = str(row.get('No_Produit', '')).strip()
                    
                    # RECHERCHEX équivalent à =SIERREUR(RECHERCHEX(S2&T2&AB2; ...); "")
                    code_promo_val = ""
                    if d_s and d_t:
                        search_key = f"{d_s}{d_t}{val_ab}"
                        code_promo_val = rabais_lookup.get(search_key, "")
                        if pd.isna(code_promo_val):
                            code_promo_val = ""
                    
                    code_promo_str = str(code_promo_val).strip()

                    cle_cmd = str(row.get('Clé_unique_détail_commande', ''))
                    if cle_cmd == 'nan' or not cle_cmd:
                        cle_cmd = ""
                        
                    cle_fact = str(row.get('Clé_unique_détail_facture', '')) if pd.notnull(row.get('Clé_unique_détail_facture', '')) else ''
                    if cle_fact == 'nan' or not cle_fact:
                        cle_fact = ""
                        
                    cle_cred_val = str(row.get(col_cred, '')) if col_cred and pd.notnull(row.get(col_cred, '')) else ''
                    date_recl = row.get('Date_Réclamée', None)
                    
                    # Règle 1 (Col B)
                    suppr_1 = "Supprimer" if pd.notnull(date_recl) and str(date_recl).strip() != "" and str(date_recl) != "NaT" else ""
                    
                    # Règle 2 / Col C
                    val_col_c = cle_cmd if (cle_cmd in cles_reclamees and cle_cmd != "") else ""
                    suppr_d = "Supprimer" if val_col_c != "" else ""
                    
                    # Règle 3 / Col E
                    cond_suppr_3 = ((cle_cred_val in cles_facture and cle_cred_val != '' and cle_cred_val != 'nan') or 
                                     (cle_fact in cles_credite and cle_fact != '' and cle_fact != 'nan'))
                    suppr_3 = "Supprimer" if cond_suppr_3 else ""
                    
                    # Règle 4 / Col F & G
                    cond_credit_remplie = (cle_fact in cles_credite and cle_fact != '' and cle_fact != 'nan')
                    val_col_f = cle_cmd if cond_credit_remplie else ""
                    suppr_g = "Supprimer" if val_col_f != "" else ""
                    
                    # Règle 5 / Col H
                    suppr_5 = "Supprimer" if qte < 0 else ""
                    
                    # Règle 6 / Col I
                    suppr_6 = ""
                    is_no_recl = (pd.isnull(date_recl) or str(date_recl).strip() == "" or str(date_recl) == "NaT")
                    if is_no_recl:
                        if sum_qte_i_group.iloc[idx] >= 0:
                            if row['_rabais_calc'] < max_rab_i_group.iloc[idx]:
                                suppr_6 = "Supprimer"
                    
                    # Règle 7 / Col J
                    suppr_7 = ""
                    val_n = row['_col_N']
                    if val_n == 0:
                        suppr_7 = "Supprimer"
                    else:
                        m_f = max_fact_n1_filled.iloc[idx]
                        if pd.notnull(row.get('Clé_unique_détail_facture')) and pd.notnull(m_f) and row.get('Clé_unique_détail_facture') < m_f:
                            suppr_7 = "Supprimer"

                    # Règle 8 / Col K
                    suppr_8 = ""
                    cond1 = (sum_qte_group.iloc[idx] > 0)
                    cond2 = group_has_date_rc.iloc[idx]
                    cond3 = (group_cred_count.iloc[idx] == 0)
                    cond4 = (montant_st < 0.99)
                    if cond1 and cond2 and cond3 and cond4:
                        suppr_8 = "Supprimer"

                    # Règle 9 / Col L
                    suppr_9 = ""
                    if code_promo_str.upper().startswith("FIL"):
                        suppr_9 = "Supprimer"

                    # Règle 10 / Col M
                    suppr_10 = "Supprimer" if montant_st < 0.99 else ""
                    
                    # Calculs financiers
                    rabais_entre_2_dates = row['_rabais_calc']
                    rabais_total = qte * montant_st
                    ecart = rabais_total - rabais_entre_2_dates
                    indicateur_tolerance = 0
                    val_col_n = row['_col_N']
                    
                    # Total Supprimer (Col A)
                    tous_criteres = [suppr_1, suppr_d, suppr_3, suppr_g, suppr_5, suppr_6, suppr_7, suppr_8, suppr_9, suppr_10]
                    suppr_total = "Supprimer" if any(c == "Supprimer" for c in tous_criteres) else ""
                    
                    # Écriture dans le classeur
                    ws_cmd.cell(row=r, column=1).value = suppr_total
                    ws_cmd.cell(row=r, column=2).value = suppr_1
                    ws_cmd.cell(row=r, column=3).value = val_col_c
                    ws_cmd.cell(row=r, column=4).value = suppr_d
                    ws_cmd.cell(row=r, column=5).value = suppr_3
                    ws_cmd.cell(row=r, column=6).value = val_col_f
                    ws_cmd.cell(row=r, column=7).value = suppr_g
                    ws_cmd.cell(row=r, column=8).value = suppr_5
                    ws_cmd.cell(row=r, column=9).value = suppr_6
                    ws_cmd.cell(row=r, column=10).value = suppr_7
                    ws_cmd.cell(row=r, column=11).value = suppr_8
                    ws_cmd.cell(row=r, column=12).value = suppr_9
                    ws_cmd.cell(row=r, column=13).value = suppr_10
                    
                    ws_cmd.cell(row=r, column=14).value = val_col_n
                    ws_cmd.cell(row=r, column=15).value = rabais_total
                    ws_cmd.cell(row=r, column=16).value = rabais_entre_2_dates
                    ws_cmd.cell(row=r, column=17).value = indicateur_tolerance
                    ws_cmd.cell(row=r, column=18).value = tolerance
                    
                    if date_deb_arr[idx]:
                        ws_cmd.cell(row=r, column=19).value = date_deb_arr[idx]
                    if date_fin_arr[idx]:
                        ws_cmd.cell(row=r, column=20).value = date_fin_arr[idx]
                        
                    ws_cmd.cell(row=r, column=21).value = code_promo_str  # Col U : RECHERCHEX appliqué
                    ws_cmd.cell(row=r, column=22).value = ecart            # Col V : Écart
            
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            
            timestamp_str = datetime.now(ZoneInfo("America/Montreal")).strftime("%Y%m%d_%H%M")
            nom_fichier = f"Rapport_Rabais_Final_v{st.session_state.version_compteur}_{timestamp_str}.xlsx"
            
            st.success(f"Traitement instantané et terminé avec succès ! ({max_row - 1} lignes traitées)")
            
            if st.download_button(
                label=f"📥 Télécharger le rapport ({nom_fichier})",
                data=output_buffer,
                file_name=nom_fichier,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ):
                st.session_state.version_compteur += 1
        else:
            st.error("Les onglets 'Rabais fournisseurs' et/ou 'Rabais entre 2 dates' sont introuvables.")
            
    except Exception as e:
        st.error(f"Une erreur s'est produite lors du traitement : {e}")
